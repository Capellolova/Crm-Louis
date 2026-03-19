import base64
import json
import os
import shutil
import sqlite3
import uuid
from contextlib import closing
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / 'data'
UPLOAD_DIR = DATA_DIR / 'uploads'
DB_PATH = DATA_DIR / 'crm.sqlite3'
SEED_XLSX = Path('/mnt/data/AFFAIRES CHAUDES LLD  .xlsx')

STATUSES = [
    'Définition du besoin',
    'Chiffrage en cours',
    'Faire proposition',
    'Proposition envoyée',
    'Relance',
    'Contre-proposition',
    'Accord client',
    'Commande en cours',
    'Gagné',
    'Perdu',
    'Stand-by',
]

PRIORITIES = ['🔥 Chaud', '⚠️ À suivre', '🧊 Froid']
TYPE_OPPS = ['Prospect', 'Client existant', 'Renouvellement', 'AO / SAD']
CATEGORIES = ['VHL', 'VUL', '3.5T', '5.2T', '10T', '16T', 'Frigo']
GAMMES = ['GL', 'VU', 'REM']
CARROSSERIES = ['FRIGO', 'HYDRAU', 'SEC', 'TRR', 'AUTRE']
ENERGIES = ['B7', 'B100', 'B100 FLEXIBLE', 'E-TECH']
CLIENT_TYPES = ['Prospect', 'Client', 'Ancien client', 'Grand compte', 'AO / public']
CLIENT_STATUS = ['Actif', 'À relancer', 'En sommeil']
POTENTIALS = ['Faible', 'Moyen', 'Fort']
CONTACT_ROLES = ['Décideur', 'Utilisateur', 'Acheteur', 'Prescripteur']
ACTIONS = [
    'Appeler', 'Relancer mail', 'Envoyer proposition', 'Refaire chiffrage',
    'Prendre RDV', 'Attendre retour client', 'Attendre validation interne',
    'Commander véhicule', 'Clôturer gagné', 'Clôturer perdu'
]
BLOCKAGES = [
    'Prix', 'Délais', 'Pas de besoin immédiat', 'Décision interne client',
    'Attente budget', 'Concurrent en place', 'Caractéristique véhicule',
    'Financement', 'Validation hiérarchique', 'Inconnu'
]


def safe_index(options, value, default=0):
    try:
        return options.index(value)
    except Exception:
        return default


def vehicle_label_parts(row_like):
    gamme = str((row_like.get('gamme') if hasattr(row_like, 'get') else '') or '').strip()
    bodywork = str((row_like.get('bodywork') if hasattr(row_like, 'get') else '') or '').strip()
    energy = str((row_like.get('energy') if hasattr(row_like, 'get') else '') or '').strip()
    parts = [p for p in [gamme, bodywork, energy] if p]
    return ' - '.join(parts) if parts else 'Véhicule non renseigné'


# ---------- DB helpers ----------
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with closing(get_conn()) as conn:
        cur = conn.cursor()
        cur.execute(
            '''CREATE TABLE IF NOT EXISTS clients (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                sector TEXT,
                account_type TEXT,
                address TEXT,
                city TEXT,
                postal_code TEXT,
                phone TEXT,
                email TEXT,
                client_status TEXT,
                potential TEXT,
                competitor TEXT,
                notes TEXT,
                contact1_name TEXT,
                contact1_role TEXT,
                contact1_title TEXT,
                contact1_phone TEXT,
                contact1_email TEXT,
                contact1_notes TEXT,
                contact2_name TEXT,
                contact2_role TEXT,
                contact2_title TEXT,
                contact2_phone TEXT,
                contact2_email TEXT,
                contact2_notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )'''
        )
        cur.execute(
            '''CREATE TABLE IF NOT EXISTS affairs (
                id TEXT PRIMARY KEY,
                client_id TEXT,
                priority TEXT,
                created_on TEXT,
                assigned_to TEXT,
                opportunity_type TEXT,
                category TEXT,
                gamme TEXT,
                ptac TEXT,
                bodywork TEXT,
                energy TEXT,
                vn_parc TEXT,
                duration_months INTEGER,
                annual_km INTEGER,
                monthly_rent REAL,
                total_estimated REAL,
                status TEXT,
                proposal_sent_on TEXT,
                next_action_date TEXT,
                next_action TEXT,
                blockage TEXT,
                competitor TEXT,
                contract_ref TEXT,
                ao_deadline TEXT,
                comments TEXT,
                last_activity_on TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(client_id) REFERENCES clients(id)
            )'''
        )
        cur.execute(
            '''CREATE TABLE IF NOT EXISTS documents (
                id TEXT PRIMARY KEY,
                affair_id TEXT NOT NULL,
                filename TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                doc_type TEXT NOT NULL,
                is_main_proposal INTEGER NOT NULL DEFAULT 0,
                uploaded_at TEXT NOT NULL,
                FOREIGN KEY(affair_id) REFERENCES affairs(id)
            )'''
        )
        conn.commit()


def now_iso():
    return datetime.now().isoformat(timespec='seconds')


def to_date_str(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if not v:
        return None
    try:
        return pd.to_datetime(v).date().isoformat()
    except Exception:
        return None


def import_seed_if_needed():
    with closing(get_conn()) as conn:
        count = conn.execute('SELECT COUNT(*) FROM affairs').fetchone()[0]
        if count > 0 or not SEED_XLSX.exists():
            return

        wb = load_workbook(SEED_XLSX, data_only=False)
        ws = wb['AFFAIRES CHAUDES']
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else '' for h in rows[0]]
        for r in rows[1:]:
            row = dict(zip(headers, r))
            client_name = str(row.get('Client ', '') or '').strip()
            if not client_name:
                continue
            client_id = slug_id('client', client_name)
            cur = conn.cursor()
            existing = cur.execute('SELECT id FROM clients WHERE id = ?', (client_id,)).fetchone()
            ts = now_iso()
            if not existing:
                cur.execute(
                    '''INSERT INTO clients (
                        id, name, sector, account_type, client_status, potential, competitor, notes, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (
                        client_id,
                        client_name,
                        str(row.get('Secteur') or ''),
                        'Client',
                        'Actif',
                        'Moyen',
                        str(row.get('Concurrent') or ''),
                        '',
                        ts,
                        ts,
                    )
                )
            affair_id = str(uuid.uuid4())
            monthly_rent = parse_float(row.get('Loyer mensuel €'))
            duration = parse_int(row.get('Durée'))
            total = monthly_rent * duration if monthly_rent and duration else None
            priority_map = {'🔥': '🔥 Chaud', '⚠️': '⚠️ À suivre', '🧊': '🧊 Froid'}
            status = normalize_status(str(row.get('Statut') or 'Définition du besoin'))
            cur.execute(
                '''INSERT INTO affairs (
                    id, client_id, priority, created_on, opportunity_type, category, gamme, ptac, bodywork,
                    energy, vn_parc, duration_months, monthly_rent, total_estimated, status, proposal_sent_on,
                    next_action_date, next_action, blockage, competitor, comments, last_activity_on, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (
                    affair_id,
                    client_id,
                    priority_map.get(str(row.get('Priorité') or '').strip(), '⚠️ À suivre'),
                    to_date_str(row.get('Date')),
                    infer_opportunity_type(status),
                    infer_category(str(row.get('Gamme') or ''), str(row.get('PTAC') or ''), str(row.get('Carrosserie') or '')),
                    str(row.get('Gamme') or ''),
                    str(row.get('PTAC') or ''),
                    str(row.get('Carrosserie') or ''),
                    str(row.get('Énergie') or ''),
                    str(row.get('VN / Parc CD') or ''),
                    duration,
                    monthly_rent,
                    total,
                    status,
                    to_date_str(row.get('Envoi proposition')),
                    to_date_str(row.get('Date prochaine action')),
                    str(row.get('Action') or ''),
                    str(row.get('Blocage principal') or ''),
                    str(row.get('Concurrent') or ''),
                    str(row.get('Commentaire') or ''),
                    to_date_str(row.get('Date prochaine action')) or to_date_str(row.get('Date')),
                    ts,
                )
            )
        conn.commit()


def slug_id(prefix, name):
    clean = ''.join(ch.lower() if ch.isalnum() else '-' for ch in name).strip('-')
    clean = '-'.join(filter(None, clean.split('-')))
    return f'{prefix}-{clean[:50]}'


def parse_float(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('€', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except Exception:
        return None


def parse_int(v):
    if v is None or v == '':
        return None
    if isinstance(v, int):
        return v
    try:
        return int(float(str(v).replace(',', '.')))
    except Exception:
        return None




def coerce_int(v, default=0):
    try:
        parsed = parse_int(v)
        return int(parsed) if parsed is not None else default
    except Exception:
        return default


def coerce_float(v, default=0.0):
    try:
        parsed = parse_float(v)
        return float(parsed) if parsed is not None else default
    except Exception:
        return default


def coerce_date(v, default=None):
    if default is None:
        default = date.today()
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v in (None, '', 'None'):
        return default
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return default


def normalize_status(status):
    mapping = {
        'Faire Proposition': 'Faire proposition',
        'Devis envoyé': 'Proposition envoyée',
        'Envoyée': 'Proposition envoyée',
        'Perdu': 'Perdu',
        'Gagné': 'Gagné',
        'Faire proposition': 'Faire proposition',
    }
    if status in STATUSES:
        return status
    return mapping.get(status, 'Définition du besoin')


def infer_opportunity_type(status):
    if status in {'Gagné', 'Commande en cours'}:
        return 'Client existant'
    return 'Prospect'


def infer_category(gamme, ptac, bodywork):
    txt = ' '.join([gamme, ptac, bodywork]).lower()
    if 'frigo' in txt or 'fgtd' in txt or 'réfrig' in txt:
        return 'Frigo'
    if '10' in txt:
        return '10T'
    if '16' in txt:
        return '16T'
    if '5.2' in txt or '5,2' in txt:
        return '5.2T'
    if '3t5' in txt or '3.5' in txt or '3,5' in txt:
        return '3.5T'
    if 'vl' in txt or 'vhl' in txt or 'gl' in txt:
        return 'VHL'
    return 'VUL'


def query_df(sql, params=()):
    with closing(get_conn()) as conn:
        return pd.read_sql_query(sql, conn, params=params)


def get_clients():
    return query_df('SELECT * FROM clients ORDER BY name')


def get_affairs():
    sql = '''
        SELECT a.*, c.name AS client_name
        FROM affairs a
        LEFT JOIN clients c ON c.id = a.client_id
        ORDER BY COALESCE(a.next_action_date, a.updated_at) ASC, a.updated_at DESC
    '''
    df = query_df(sql)
    if not df.empty:
        for col in ['created_on', 'proposal_sent_on', 'next_action_date', 'ao_deadline', 'last_activity_on']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.date
        df['days_without_activity'] = (pd.Timestamp.today().normalize() - pd.to_datetime(df['last_activity_on'], errors='coerce')).dt.days
        df['retard_action'] = pd.to_datetime(df['next_action_date'], errors='coerce').dt.date < date.today()
        df['vehicle_label'] = df.apply(vehicle_label_parts, axis=1)
    return df


def get_documents(affair_id):
    return query_df('SELECT * FROM documents WHERE affair_id = ? ORDER BY uploaded_at DESC', (affair_id,))


# ---------- CRUD ----------
def upsert_client(data):
    ts = now_iso()
    client_id = data.get('id') or str(uuid.uuid4())
    with closing(get_conn()) as conn:
        cur = conn.cursor()
        exists = cur.execute('SELECT 1 FROM clients WHERE id = ?', (client_id,)).fetchone()
        payload = (
            client_id, data['name'], data.get('sector'), data.get('account_type'), data.get('address'), data.get('city'),
            data.get('postal_code'), data.get('phone'), data.get('email'), data.get('client_status'), data.get('potential'),
            data.get('competitor'), data.get('notes'), data.get('contact1_name'), data.get('contact1_role'),
            data.get('contact1_title'), data.get('contact1_phone'), data.get('contact1_email'), data.get('contact1_notes'),
            data.get('contact2_name'), data.get('contact2_role'), data.get('contact2_title'), data.get('contact2_phone'),
            data.get('contact2_email'), data.get('contact2_notes')
        )
        if exists:
            cur.execute('''UPDATE clients SET
                name=?, sector=?, account_type=?, address=?, city=?, postal_code=?, phone=?, email=?, client_status=?,
                potential=?, competitor=?, notes=?, contact1_name=?, contact1_role=?, contact1_title=?, contact1_phone=?,
                contact1_email=?, contact1_notes=?, contact2_name=?, contact2_role=?, contact2_title=?, contact2_phone=?,
                contact2_email=?, contact2_notes=?, updated_at=? WHERE id=?''',
                payload[1:] + (ts, client_id)
            )
        else:
            cur.execute('''INSERT INTO clients (
                id, name, sector, account_type, address, city, postal_code, phone, email, client_status, potential,
                competitor, notes, contact1_name, contact1_role, contact1_title, contact1_phone, contact1_email,
                contact1_notes, contact2_name, contact2_role, contact2_title, contact2_phone, contact2_email,
                contact2_notes, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                payload + (ts, ts)
            )
        conn.commit()
    return client_id


def upsert_affair(data):
    ts = now_iso()
    affair_id = data.get('id') or str(uuid.uuid4())
    duration = parse_int(data.get('duration_months'))
    monthly = parse_float(data.get('monthly_rent'))
    total = monthly * duration if monthly and duration else None
    with closing(get_conn()) as conn:
        cur = conn.cursor()
        exists = cur.execute('SELECT 1 FROM affairs WHERE id = ?', (affair_id,)).fetchone()
        payload = (
            affair_id, data.get('client_id'), data.get('priority'), to_date_str(data.get('created_on')),
            data.get('assigned_to'), data.get('opportunity_type'), None, data.get('gamme'),
            data.get('ptac'), data.get('bodywork'), data.get('energy'), data.get('vn_parc'), duration,
            parse_int(data.get('annual_km')), monthly, total, data.get('status'), to_date_str(data.get('proposal_sent_on')),
            to_date_str(data.get('next_action_date')), data.get('next_action'), data.get('blockage'), data.get('competitor'),
            data.get('contract_ref'), to_date_str(data.get('ao_deadline')), data.get('comments'),
            to_date_str(data.get('last_activity_on')) or date.today().isoformat(), ts
        )
        if exists:
            cur.execute('''UPDATE affairs SET
                client_id=?, priority=?, created_on=?, assigned_to=?, opportunity_type=?, category=?, gamme=?, ptac=?, bodywork=?,
                energy=?, vn_parc=?, duration_months=?, annual_km=?, monthly_rent=?, total_estimated=?, status=?, proposal_sent_on=?,
                next_action_date=?, next_action=?, blockage=?, competitor=?, contract_ref=?, ao_deadline=?, comments=?,
                last_activity_on=?, updated_at=? WHERE id=?''',
                payload[1:] + (affair_id,)
            )
        else:
            cur.execute('''INSERT INTO affairs (
                id, client_id, priority, created_on, assigned_to, opportunity_type, category, gamme, ptac, bodywork,
                energy, vn_parc, duration_months, annual_km, monthly_rent, total_estimated, status, proposal_sent_on,
                next_action_date, next_action, blockage, competitor, contract_ref, ao_deadline, comments, last_activity_on,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', payload)
        conn.commit()
    return affair_id


def save_document(affair_id, uploaded_file, doc_type, is_main_proposal):
    ext = Path(uploaded_file.name).suffix or '.bin'
    doc_id = str(uuid.uuid4())
    stored_name = f'{doc_id}{ext}'
    stored_path = UPLOAD_DIR / stored_name
    with open(stored_path, 'wb') as f:
        f.write(uploaded_file.getbuffer())
    with closing(get_conn()) as conn:
        cur = conn.cursor()
        if is_main_proposal:
            cur.execute('UPDATE documents SET is_main_proposal = 0 WHERE affair_id = ?', (affair_id,))
        cur.execute(
            'INSERT INTO documents (id, affair_id, filename, stored_path, doc_type, is_main_proposal, uploaded_at) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (doc_id, affair_id, uploaded_file.name, str(stored_path), doc_type, 1 if is_main_proposal else 0, now_iso())
        )
        conn.commit()


def update_affair_after_action(affair_id, status=None, next_action_date=None, log_line=None):
    with closing(get_conn()) as conn:
        cur = conn.cursor()
        row = cur.execute('SELECT comments FROM affairs WHERE id = ?', (affair_id,)).fetchone()
        comments = row['comments'] if row else ''
        if log_line:
            comments = f"{datetime.now().strftime('%d/%m/%Y')} : {log_line}\n{comments or ''}".strip()
        cur.execute('''UPDATE affairs SET status = COALESCE(?, status), next_action_date = COALESCE(?, next_action_date),
                       last_activity_on = ?, comments = ?, updated_at = ? WHERE id = ?''',
                    (status, to_date_str(next_action_date), date.today().isoformat(), comments, now_iso(), affair_id))
        conn.commit()


# ---------- email draft ----------
def get_main_proposal_doc(affair_id):
    with closing(get_conn()) as conn:
        row = conn.execute('SELECT * FROM documents WHERE affair_id = ? AND is_main_proposal = 1 ORDER BY uploaded_at DESC LIMIT 1', (affair_id,)).fetchone()
        return dict(row) if row else None


def build_relance_email(affair_row, client_row):
    contact_name = client_row.get('contact1_name') or client_row.get('contact2_name') or client_row.get('name')
    firstname = (contact_name.split()[0] if contact_name else '').strip()
    vehicle_label = vehicle_label_parts(affair_row)
    subject = f"Relance – Proposition {vehicle_label} – {client_row.get('name')}"
    body = f"""Bonjour {firstname or ''},

Je me permets de revenir vers vous suite à l’envoi de notre proposition concernant {vehicle_label}.

Avez-vous eu l’occasion d’en prendre connaissance ?

Je reste bien entendu à votre disposition pour en échanger ou ajuster certains points si nécessaire.

Je vous remets la proposition en pièce jointe pour plus de facilité.

Dans l’attente de votre retour,

Bonne journée,

Louis Capello
Clovis / Locatrucks
""".replace('Bonjour ,', 'Bonjour,')
    return subject, body


def create_eml_draft(affair_row, client_row, output_dir: Path):
    output_dir.mkdir(exist_ok=True, parents=True)
    subject, body = build_relance_email(affair_row, client_row)
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = 'louis@example.com'
    recipient = client_row.get('contact1_email') or client_row.get('contact2_email') or client_row.get('email') or 'destinataire@example.com'
    msg['To'] = recipient
    msg.set_content(body)

    main_doc = get_main_proposal_doc(affair_row['id'])
    if main_doc and Path(main_doc['stored_path']).exists():
        with open(main_doc['stored_path'], 'rb') as f:
            data = f.read()
        maintype = 'application'
        subtype = 'pdf' if main_doc['filename'].lower().endswith('.pdf') else 'octet-stream'
        msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=main_doc['filename'])

    eml_path = output_dir / f"relance_{affair_row['id']}.eml"
    with open(eml_path, 'wb') as f:
        f.write(bytes(msg))
    return eml_path, bool(main_doc)


# ---------- UI helpers ----------
def metric_card(label, value, help_text=None):
    st.metric(label, value, help=help_text)


def get_alerts(affairs_df):
    alerts = []
    if affairs_df.empty:
        return alerts
    today = date.today()
    for _, row in affairs_df.iterrows():
        if pd.notna(row.get('ao_deadline')) and row['ao_deadline']:
            delta = (row['ao_deadline'] - today).days
            if 0 <= delta <= 3:
                alerts.append((1, f"AO à rendre dans {delta} jour(s) — {row['client_name']}", row['id']))
            elif delta < 0:
                alerts.append((0, f"Deadline AO dépassée — {row['client_name']}", row['id']))
        if row.get('priority') == '🔥 Chaud' and (row.get('days_without_activity') or 0) >= 2:
            alerts.append((2, f"Affaire chaude inactive — {row['client_name']}", row['id']))
        next_action = row.get('next_action_date')
        if pd.notna(next_action) and next_action:
            if next_action < today:
                alerts.append((1, f"Action en retard — {row['client_name']}", row['id']))
            elif next_action == today:
                alerts.append((3, f"Action prévue aujourd'hui — {row['client_name']}", row['id']))
    return sorted(alerts, key=lambda x: x[0])


def load_row_as_dict(df, row_id):
    if df.empty:
        return {}
    match = df[df['id'] == row_id]
    if match.empty:
        return {}
    rec = match.iloc[0].to_dict()
    for k, v in list(rec.items()):
        if pd.isna(v):
            rec[k] = None
    return rec

def render_compact_cards(df, fields, title_field=None):
    if df.empty:
        st.info('Aucune donnée à afficher.')
        return
    for _, row in df.iterrows():
        title = f"**{row.get(title_field) or '-'}**\n\n" if title_field else ''
        lines = []
        for field, label in fields:
            value = row.get(field)
            if pd.isna(value) if hasattr(pd, 'isna') else False:
                value = None
            if value in [None, '', 'NaT']:
                value = '-'
            lines.append(f"**{label}** : {value}")
        st.markdown(title + '\n\n'.join(lines))
        st.markdown('---')



# ---------- pages ----------
def page_dashboard():
    st.title('Tableau de bord')
    affairs = get_affairs()
    clients = get_clients()

    signed = affairs[affairs['status'] == 'Gagné']['total_estimated'].fillna(0).sum() if not affairs.empty else 0
    pipeline = affairs[~affairs['status'].isin(['Gagné', 'Perdu'])]['total_estimated'].fillna(0).sum() if not affairs.empty else 0
    hot = int(((affairs['priority'] == '🔥 Chaud') & (~affairs['status'].isin(['Gagné', 'Perdu']))).sum()) if not affairs.empty else 0
    today_actions = int((affairs['next_action_date'] == date.today()).sum()) if not affairs.empty else 0
    overdue = int((affairs['retard_action'] == True).sum()) if not affairs.empty else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        metric_card('CA signé', f"{signed:,.0f} €".replace(',', ' '))
    with c2:
        metric_card('CA en cours', f"{pipeline:,.0f} €".replace(',', ' '))
    with c3:
        metric_card('Affaires chaudes', hot)
    with c4:
        metric_card("Actions aujourd'hui", today_actions)
    with c5:
        metric_card('Retards', overdue)

    st.subheader('Plan de bataille')
    alerts = get_alerts(affairs)
    if alerts:
        for _, msg, _ in alerts[:8]:
            st.warning(msg)
    else:
        st.success('RAS : le pipe est propre.')

    left, right = st.columns([1.2, 1])
    with left:
        st.subheader('Pipeline')
        if affairs.empty:
            st.info('Aucune affaire pour le moment.')
        else:
            pipe = affairs.groupby('status', dropna=False).agg(nombre=('id', 'count'), montant_total=('total_estimated', 'sum')).reset_index()
            pipe = pipe.rename(columns={'status': 'Statut', 'montant_total': 'Montant total'})
            tab_table, tab_mobile = st.tabs(['Tableau', 'Mobile'])
            with tab_table:
                st.dataframe(pipe, use_container_width=True, hide_index=True)
            with tab_mobile:
                render_compact_cards(pipe, [('nombre', 'Nombre'), ('Montant total', 'Montant total')], 'Statut')

        st.subheader('Top opportunités')
        if not affairs.empty:
            top = affairs.sort_values(['priority', 'total_estimated'], ascending=[True, False]).head(5)
            top_display = top[['client_name', 'vehicle_label', 'status', 'total_estimated', 'next_action_date']].rename(columns={
                'client_name': 'Client', 'vehicle_label': 'Véhicule', 'status': 'Statut', 'total_estimated': 'Montant estimé', 'next_action_date': 'Prochaine action'
            })
            tab_table, tab_mobile = st.tabs(['Tableau', 'Mobile'])
            with tab_table:
                st.dataframe(top_display, use_container_width=True, hide_index=True)
            with tab_mobile:
                render_compact_cards(top_display, [('Véhicule', 'Véhicule'), ('Statut', 'Statut'), ('Montant estimé', 'Montant'), ('Prochaine action', 'Prochaine action')], 'Client')
    with right:
        st.subheader('Affaires qui dorment')
        if not affairs.empty:
            sleep = affairs[affairs['days_without_activity'].fillna(0) >= 10]
            if sleep.empty:
                st.success('Aucune affaire en train de mourir. Miracle commercial.')
            else:
                sleep_display = sleep[['client_name', 'status', 'days_without_activity', 'next_action_date']].rename(columns={
                    'client_name': 'Client', 'status': 'Statut', 'days_without_activity': 'Jours sans activité', 'next_action_date': 'Prochaine action'
                })
                tab_table, tab_mobile = st.tabs(['Tableau', 'Mobile'])
                with tab_table:
                    st.dataframe(sleep_display, use_container_width=True, hide_index=True)
                with tab_mobile:
                    render_compact_cards(sleep_display, [('Statut', 'Statut'), ('Jours sans activité', 'Jours sans activité'), ('Prochaine action', 'Prochaine action')], 'Client')

        st.subheader('Perf AO / SAD')
        ao = affairs[affairs['opportunity_type'] == 'AO / SAD'] if not affairs.empty else pd.DataFrame()
        if ao.empty:
            st.info('Pas encore d’AO saisis.')
        else:
            ao_won = int((ao['status'] == 'Gagné').sum())
            ao_total = len(ao)
            ao_rate = round((ao_won / ao_total) * 100, 1) if ao_total else 0
            st.write(f"AO reçus : **{ao_total}**")
            st.write(f"AO gagnés : **{ao_won}**")
            st.write(f"Taux de réussite : **{ao_rate}%**")


def page_clients():
    st.title('Clients')
    clients = get_clients()
    affairs = get_affairs()

    with st.expander('➕ Nouveau / modifier client', expanded=False):
        selected_id = st.selectbox('Client à modifier', ['Nouveau client'] + clients['name'].tolist() if not clients.empty else ['Nouveau client'])
        current = {}
        if selected_id != 'Nouveau client':
            current = load_row_as_dict(clients, clients[clients['name'] == selected_id].iloc[0]['id'])
        with st.form('client_form'):
            c1, c2 = st.columns(2)
            with c1:
                name = st.text_input('Nom du client', value=current.get('name') or '')
                sector = st.text_input('Secteur / département', value=current.get('sector') or '')
                account_type = st.selectbox('Type de compte', CLIENT_TYPES, index=CLIENT_TYPES.index(current.get('account_type')) if current.get('account_type') in CLIENT_TYPES else 0)
                client_status = st.selectbox('Statut client', CLIENT_STATUS, index=CLIENT_STATUS.index(current.get('client_status')) if current.get('client_status') in CLIENT_STATUS else 0)
                potential = st.selectbox('Potentiel', POTENTIALS, index=POTENTIALS.index(current.get('potential')) if current.get('potential') in POTENTIALS else 1)
                competitor = st.text_input('Concurrent en place', value=current.get('competitor') or '')
            with c2:
                address = st.text_input('Adresse', value=current.get('address') or '')
                city = st.text_input('Ville', value=current.get('city') or '')
                postal_code = st.text_input('Code postal', value=current.get('postal_code') or '')
                phone = st.text_input('Téléphone standard', value=current.get('phone') or '')
                email = st.text_input('Email générique', value=current.get('email') or '')
            st.markdown('**Contact clé 1**')
            x1, x2, x3 = st.columns(3)
            with x1:
                contact1_name = st.text_input('Nom / prénom contact 1', value=current.get('contact1_name') or '')
                contact1_title = st.text_input('Fonction contact 1', value=current.get('contact1_title') or '')
            with x2:
                contact1_phone = st.text_input('Téléphone contact 1', value=current.get('contact1_phone') or '')
                contact1_email = st.text_input('Email contact 1', value=current.get('contact1_email') or '')
            with x3:
                contact1_role = st.selectbox('Rôle contact 1', CONTACT_ROLES, index=CONTACT_ROLES.index(current.get('contact1_role')) if current.get('contact1_role') in CONTACT_ROLES else 0)
                contact1_notes = st.text_input('Notes contact 1', value=current.get('contact1_notes') or '')
            st.markdown('**Contact clé 2**')
            y1, y2, y3 = st.columns(3)
            with y1:
                contact2_name = st.text_input('Nom / prénom contact 2', value=current.get('contact2_name') or '')
                contact2_title = st.text_input('Fonction contact 2', value=current.get('contact2_title') or '')
            with y2:
                contact2_phone = st.text_input('Téléphone contact 2', value=current.get('contact2_phone') or '')
                contact2_email = st.text_input('Email contact 2', value=current.get('contact2_email') or '')
            with y3:
                contact2_role = st.selectbox('Rôle contact 2', CONTACT_ROLES, index=CONTACT_ROLES.index(current.get('contact2_role')) if current.get('contact2_role') in CONTACT_ROLES else 1)
                contact2_notes = st.text_input('Notes contact 2', value=current.get('contact2_notes') or '')
            notes = st.text_area('Notes générales', value=current.get('notes') or '')
            submitted = st.form_submit_button('Enregistrer le client')
            if submitted:
                if not name.strip():
                    st.error('Le nom du client est obligatoire.')
                else:
                    cid = upsert_client({
                        'id': current.get('id'), 'name': name.strip(), 'sector': sector, 'account_type': account_type,
                        'address': address, 'city': city, 'postal_code': postal_code, 'phone': phone, 'email': email,
                        'client_status': client_status, 'potential': potential, 'competitor': competitor, 'notes': notes,
                        'contact1_name': contact1_name, 'contact1_role': contact1_role, 'contact1_title': contact1_title,
                        'contact1_phone': contact1_phone, 'contact1_email': contact1_email, 'contact1_notes': contact1_notes,
                        'contact2_name': contact2_name, 'contact2_role': contact2_role, 'contact2_title': contact2_title,
                        'contact2_phone': contact2_phone, 'contact2_email': contact2_email, 'contact2_notes': contact2_notes,
                    })
                    st.success(f'Client enregistré : {name}')
                    st.rerun()

    st.subheader('Base clients')
    if clients.empty:
        st.info('Aucun client pour le moment.')
    else:
        display = clients[['name', 'sector', 'account_type', 'client_status', 'potential', 'contact1_name', 'contact2_name']].copy()
        display = display.rename(columns={'name':'Client','sector':'Secteur','account_type':'Type de compte','client_status':'Statut client','potential':'Potentiel','contact1_name':'Contact 1','contact2_name':'Contact 2'})
        tab_table, tab_mobile = st.tabs(['Tableau', 'Mobile'])
        with tab_table:
            st.dataframe(display, use_container_width=True, hide_index=True)
        with tab_mobile:
            render_compact_cards(display, [('Secteur','Secteur'),('Type de compte','Type de compte'),('Statut client','Statut'),('Potentiel','Potentiel'),('Contact 1','Contact 1'),('Contact 2','Contact 2')], 'Client')

        client_name = st.selectbox('Ouvrir une fiche client', clients['name'].tolist())
        client = clients[clients['name'] == client_name].iloc[0].to_dict()
        linked = affairs[affairs['client_id'] == client['id']] if not affairs.empty else pd.DataFrame()
        st.markdown(f"### {client['name']}")
        a, b, c = st.columns(3)
        a.write(f"**Statut** : {client.get('client_status') or '-'}")
        b.write(f"**Potentiel** : {client.get('potential') or '-'}")
        c.write(f"**Concurrent** : {client.get('competitor') or '-'}")
        st.write(f"**Contact 1** : {client.get('contact1_name') or '-'} — {client.get('contact1_title') or '-'} — {client.get('contact1_email') or '-'}")
        st.write(f"**Contact 2** : {client.get('contact2_name') or '-'} — {client.get('contact2_title') or '-'} — {client.get('contact2_email') or '-'}")
        st.write(f"**Notes** : {client.get('notes') or '-'}")
        st.markdown('**Affaires liées**')
        if linked.empty:
            st.info('Aucune affaire liée à ce client.')
        else:
            linked_display = linked[['vehicle_label', 'status', 'priority', 'monthly_rent', 'next_action_date']].rename(columns={'vehicle_label':'Véhicule','status':'Statut','priority':'Priorité','monthly_rent':'Loyer mensuel','next_action_date':'Prochaine action'})
            st.dataframe(linked_display, use_container_width=True, hide_index=True)


def page_affairs():
    st.title('Affaires')
    clients = get_clients()
    affairs = get_affairs()

    with st.expander('➕ Nouvelle / modifier affaire', expanded=False):
        name_options = ['Nouvelle affaire'] + ([f"{r['client_name']} — {r['status']} — {r['vehicle_label']}" for _, r in affairs.iterrows()] if not affairs.empty else [])
        selected = st.selectbox('Affaire à modifier', name_options)
        current = {}
        if selected != 'Nouvelle affaire' and not affairs.empty:
            idx = name_options.index(selected) - 1
            current = affairs.iloc[idx].to_dict()
        with st.form('affair_form'):
            c1, c2, c3 = st.columns(3)
            with c1:
                client_options = clients['name'].tolist() if not clients.empty else ['Aucun client disponible']
                client_idx = 0
                if not clients.empty and current.get('client_id') in set(clients['id']):
                    match_positions = clients.index[clients['id'] == current.get('client_id')].tolist()
                    if match_positions:
                        client_idx = int(match_positions[0])
                client_name = st.selectbox('Client lié', client_options, index=client_idx, disabled=clients.empty)
                priority = st.selectbox('Priorité', PRIORITIES, index=safe_index(PRIORITIES, current.get('priority'), 1))
                created_on = st.date_input('Date de création', value=coerce_date(current.get('created_on')))
                assigned_to = st.text_input('Commercial assigné', value=current.get('assigned_to') or 'Louis')
                opportunity_type = st.selectbox('Type opportunité', TYPE_OPPS, index=safe_index(TYPE_OPPS, current.get('opportunity_type'), 0))
            with c2:
                gamme = st.selectbox('Gamme', GAMMES, index=safe_index(GAMMES, current.get('gamme'), 0))
                ptac = st.text_input('PTAC', value=current.get('ptac') or '')
                bodywork = st.selectbox('Carrosserie', CARROSSERIES, index=safe_index(CARROSSERIES, current.get('bodywork'), 0))
                energy = st.selectbox('Énergie', ENERGIES, index=safe_index(ENERGIES, current.get('energy'), 0))
            with c3:
                vn_parc = st.text_input('VN / Parc', value=current.get('vn_parc') or '')
                duration_months = st.number_input('Durée (mois)', min_value=0, step=1, value=coerce_int(current.get('duration_months'), 0))
                annual_km = st.number_input('Km/an', min_value=0, step=1000, value=coerce_int(current.get('annual_km'), 0))
                monthly_rent = st.number_input('Loyer mensuel €', min_value=0.0, step=10.0, value=coerce_float(current.get('monthly_rent'), 0.0))
            s1, s2, s3 = st.columns(3)
            with s1:
                status = st.selectbox('Statut', STATUSES, index=STATUSES.index(current.get('status')) if current.get('status') in STATUSES else 0)
                proposal_sent_on = st.date_input('Date envoi proposition', value=coerce_date(current.get('proposal_sent_on')))
                next_action_date = st.date_input('Date prochaine action', value=coerce_date(current.get('next_action_date')))
            with s2:
                next_action = st.selectbox('Action suivante', ACTIONS, index=ACTIONS.index(current.get('next_action')) if current.get('next_action') in ACTIONS else 0)
                blockage = st.selectbox('Blocage principal', BLOCKAGES, index=BLOCKAGES.index(current.get('blockage')) if current.get('blockage') in BLOCKAGES else 0)
                competitor = st.text_input('Concurrent', value=current.get('competitor') or '')
            with s3:
                contract_ref = st.text_input('Contrat / BDC', value=current.get('contract_ref') or '')
                ao_deadline = st.date_input('Deadline AO', value=coerce_date(current.get('ao_deadline')))
            comments = st.text_area('Commentaires', value=current.get('comments') or '')
            submitted = st.form_submit_button('Enregistrer l’affaire')
            if submitted:
                client_id = clients[clients['name'] == client_name].iloc[0]['id'] if not clients.empty else None
                upsert_affair({
                    'id': current.get('id'), 'client_id': client_id, 'priority': priority, 'created_on': created_on,
                    'assigned_to': assigned_to, 'opportunity_type': opportunity_type,
                    'gamme': gamme, 'ptac': ptac, 'bodywork': bodywork, 'energy': energy, 'vn_parc': vn_parc,
                    'duration_months': duration_months, 'annual_km': annual_km, 'monthly_rent': monthly_rent,
                    'status': status, 'proposal_sent_on': proposal_sent_on, 'next_action_date': next_action_date,
                    'next_action': next_action, 'blockage': blockage, 'competitor': competitor,
                    'contract_ref': contract_ref, 'ao_deadline': ao_deadline, 'comments': comments,
                    'last_activity_on': date.today(),
                })
                st.success('Affaire enregistrée.')
                st.rerun()

    st.subheader('Filtres')
    f1, f2, f3 = st.columns(3)
    with f1:
        status_filter = st.multiselect('Statut', STATUSES)
    with f2:
        prio_filter = st.multiselect('Priorité', PRIORITIES)
    with f3:
        type_filter = st.multiselect('Type', TYPE_OPPS)

    filtered = affairs.copy()
    if not filtered.empty:
        if status_filter:
            filtered = filtered[filtered['status'].isin(status_filter)]
        if prio_filter:
            filtered = filtered[filtered['priority'].isin(prio_filter)]
        if type_filter:
            filtered = filtered[filtered['opportunity_type'].isin(type_filter)]

    filtered_display = filtered[['client_name', 'priority', 'opportunity_type', 'vehicle_label', 'status', 'monthly_rent', 'next_action_date', 'competitor', 'blockage']].rename(columns={
        'client_name':'Client','priority':'Priorité','opportunity_type':'Type opportunité','vehicle_label':'Véhicule','status':'Statut','monthly_rent':'Loyer mensuel','next_action_date':'Prochaine action','competitor':'Concurrent','blockage':'Blocage'
    })
    tab_table, tab_mobile = st.tabs(['Tableau', 'Mobile'])
    with tab_table:
        st.dataframe(filtered_display, use_container_width=True, hide_index=True)
    with tab_mobile:
        render_compact_cards(filtered_display, [('Priorité','Priorité'),('Type opportunité','Type'),('Véhicule','Véhicule'),('Statut','Statut'),('Loyer mensuel','Loyer'),('Prochaine action','Prochaine action'),('Concurrent','Concurrent'),('Blocage','Blocage')], 'Client')

    st.subheader('Fiche affaire')
    if affairs.empty:
        st.info('Aucune affaire à afficher.')
        return

    select_map = {f"{r['client_name']} — {r['status']} — {r['vehicle_label']}": r['id'] for _, r in affairs.iterrows()}
    selected_label = st.selectbox('Ouvrir une affaire', list(select_map.keys()))
    affair = load_row_as_dict(affairs, select_map[selected_label])
    client = load_row_as_dict(clients, affair['client_id'])

    a1, a2, a3, a4 = st.columns(4)
    a1.metric('Client', affair.get('client_name') or '-')
    a2.metric('Statut', affair.get('status') or '-')
    a3.metric('Priorité', affair.get('priority') or '-')
    a4.metric('Montant', f"{(affair.get('total_estimated') or 0):,.0f} €".replace(',', ' '))

    st.write(f"**Action suivante** : {affair.get('next_action') or '-'}")
    st.write(f"**Date prochaine action** : {affair.get('next_action_date') or '-'}")
    st.write(f"**Blocage** : {affair.get('blockage') or '-'}")
    st.write(f"**Commentaires** : {affair.get('comments') or '-'}")

    b1, b2, b3, b4 = st.columns(4)
    with b1:
        if st.button('📞 Appeler', use_container_width=True):
            st.info(f"Appelle {client.get('contact1_phone') or client.get('phone') or 'le client'}")
    with b2:
        if st.button('✉️ Relancer', use_container_width=True):
            eml_path, has_attachment = create_eml_draft(affair, client, DATA_DIR / 'drafts')
            update_affair_after_action(affair['id'], status='Relance', next_action_date=date.today() + timedelta(days=3), log_line='Relance préparée depuis l’app')
            st.success('Brouillon .eml généré.')
            if not has_attachment:
                st.warning('Aucune proposition principale en PDF trouvée. Le brouillon est créé, mais sans PJ.')
            with open(eml_path, 'rb') as f:
                st.download_button('Télécharger le brouillon email (.eml)', data=f, file_name=eml_path.name, mime='message/rfc822', use_container_width=True)
    with b3:
        if st.button('📅 Planifier', use_container_width=True):
            update_affair_after_action(affair['id'], next_action_date=date.today() + timedelta(days=2), log_line='Action replanifiée')
            st.success('Action replanifiée à J+2.')
    with b4:
        st.write('')

    st.markdown('**Documents liés**')
    docs = get_documents(affair['id'])
    if docs.empty:
        st.info('Aucun document lié.')
    else:
        st.dataframe(docs[['filename', 'doc_type', 'is_main_proposal', 'uploaded_at']], use_container_width=True, hide_index=True)
        for _, doc in docs.iterrows():
            path = Path(doc['stored_path'])
            if path.exists():
                with open(path, 'rb') as f:
                    st.download_button(f"Télécharger {doc['filename']}", data=f, file_name=doc['filename'], use_container_width=False)

    with st.form('upload_doc_form'):
        uploaded = st.file_uploader('Ajouter un document à cette affaire', type=['pdf', 'docx', 'xlsx', 'png', 'jpg'])
        doc_type = st.selectbox('Type de document', ['Proposition commerciale', 'Devis', 'Bon de commande', 'Autre'])
        is_main = st.checkbox('Définir comme proposition principale')
        submit_doc = st.form_submit_button('Ajouter le document')
        if submit_doc:
            if uploaded is None:
                st.error('Choisis un fichier avant de valider.')
            else:
                save_document(affair['id'], uploaded, doc_type, is_main)
                update_affair_after_action(affair['id'], log_line=f'Document ajouté : {uploaded.name}')
                st.success('Document ajouté.')
                st.rerun()


def page_today():
    st.title('Aujourd’hui')
    affairs = get_affairs()
    if affairs.empty:
        st.info('Aucune affaire pour le moment.')
        return
    today_df = affairs[(affairs['next_action_date'] == date.today()) | (affairs['retard_action'] == True)].copy()
    ao_df = affairs[(pd.to_datetime(affairs['ao_deadline'], errors='coerce').dt.date >= date.today()) & (pd.to_datetime(affairs['ao_deadline'], errors='coerce').dt.date <= date.today() + timedelta(days=3))].copy()
    hot_df = affairs[(affairs['priority'] == '🔥 Chaud') & (~affairs['status'].isin(['Gagné', 'Perdu'])) & (affairs['days_without_activity'].fillna(0) >= 2)].copy()

    st.subheader("À faire aujourd'hui / en retard")
    today_display = today_df[['client_name', 'status', 'next_action', 'next_action_date', 'priority']].rename(columns={'client_name':'Client','status':'Statut','next_action':'Action','next_action_date':'Date','priority':'Priorité'})
    tab_table, tab_mobile = st.tabs(['Tableau', 'Mobile'])
    with tab_table:
        st.dataframe(today_display, use_container_width=True, hide_index=True)
    with tab_mobile:
        render_compact_cards(today_display, [('Statut','Statut'),('Action','Action'),('Date','Date'),('Priorité','Priorité')], 'Client')
    st.subheader('AO à échéance proche')
    ao_display = ao_df[['client_name', 'vehicle_label', 'ao_deadline', 'status']].rename(columns={'client_name':'Client','vehicle_label':'Véhicule','ao_deadline':'Échéance AO','status':'Statut'})
    tab_table2, tab_mobile2 = st.tabs(['Tableau', 'Mobile'])
    with tab_table2:
        st.dataframe(ao_display, use_container_width=True, hide_index=True)
    with tab_mobile2:
        render_compact_cards(ao_display, [('Véhicule','Véhicule'),('Échéance AO','Échéance'),('Statut','Statut')], 'Client')
    st.subheader('Affaires chaudes inactives')
    hot_display = hot_df[['client_name', 'status', 'days_without_activity', 'monthly_rent']].rename(columns={'client_name':'Client','status':'Statut','days_without_activity':'Jours sans activité','monthly_rent':'Loyer mensuel'})
    tab_table3, tab_mobile3 = st.tabs(['Tableau', 'Mobile'])
    with tab_table3:
        st.dataframe(hot_display, use_container_width=True, hide_index=True)
    with tab_mobile3:
        render_compact_cards(hot_display, [('Statut','Statut'),('Jours sans activité','Jours sans activité'),('Loyer mensuel','Loyer')], 'Client')


def page_stats():
    st.title('Statistiques & pilotage')
    affairs = get_affairs()
    if affairs.empty:
        st.info('Pas encore de stats : saisis quelques affaires et ça commencera à parler.')
        return
    total = len(affairs)
    won = int((affairs['status'] == 'Gagné').sum())
    lost = int((affairs['status'] == 'Perdu').sum())
    rate = round((won / total) * 100, 1) if total else 0
    avg_ticket = affairs['total_estimated'].fillna(0).mean()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric('Affaires', total)
    c2.metric('Gagnées', won)
    c3.metric('Perdues', lost)
    c4.metric('Taux de transformation', f'{rate}%')
    st.metric('CA moyen par affaire', f"{avg_ticket:,.0f} €".replace(',', ' '))

    st.subheader('Répartition par véhicule')
    by_cat = affairs.groupby('vehicle_label', dropna=False).agg(nombre=('id', 'count'), montant_total=('total_estimated', 'sum')).reset_index()
    by_cat = by_cat.rename(columns={'vehicle_label': 'Véhicule', 'montant_total': 'Montant total'})
    st.dataframe(by_cat, use_container_width=True, hide_index=True)
    st.subheader('Concurrence')
    by_comp = affairs.fillna({'competitor': 'Non renseigné'}).groupby('competitor').agg(nombre=('id', 'count')).reset_index().sort_values('nombre', ascending=False)
    by_comp = by_comp.rename(columns={'competitor': 'Concurrent'})
    st.dataframe(by_comp, use_container_width=True, hide_index=True)
    st.subheader('Blocages')
    by_block = affairs.fillna({'blockage': 'Non renseigné'}).groupby('blockage').agg(nombre=('id', 'count')).reset_index().sort_values('nombre', ascending=False)
    by_block = by_block.rename(columns={'blockage': 'Blocage'})
    st.dataframe(by_block, use_container_width=True, hide_index=True)


def page_quick_add():
    st.title('Ajout rapide')
    st.caption('Créer une affaire ou un client en moins de 30 secondes. Pas de blabla, pas de tunnel administratif.')
    tab1, tab2 = st.tabs(['Nouvelle affaire', 'Nouveau client'])
    with tab1:
        clients = get_clients()
        with st.form('quick_affair'):
            c1, c2 = st.columns(2)
            with c1:
                client_name = st.selectbox('Client', clients['name'].tolist() if not clients.empty else ['Aucun client disponible'], disabled=clients.empty)
                gamme = st.selectbox('Gamme', GAMMES)
                status = st.selectbox('Statut', STATUSES, index=2)
            with c2:
                monthly_rent = st.number_input('Loyer €', min_value=0.0, step=10.0)
                next_action = st.selectbox('Action', ACTIONS)
                next_action_date = st.date_input('Quand ?', value=date.today())
            submit = st.form_submit_button('Créer')
            if submit:
                if clients.empty:
                    st.error('Crée d’abord un client.')
                else:
                    client_id = clients[clients['name'] == client_name].iloc[0]['id']
                    upsert_affair({
                        'client_id': client_id, 'priority': '⚠️ À suivre', 'created_on': date.today(), 'assigned_to': 'Louis',
                        'opportunity_type': 'Prospect', 'gamme': gamme, 'status': status, 'monthly_rent': monthly_rent,
                        'next_action': next_action, 'next_action_date': next_action_date, 'last_activity_on': date.today(),
                    })
                    st.success('Affaire créée.')
    with tab2:
        with st.form('quick_client'):
            name = st.text_input('Nom du client')
            sector = st.text_input('Secteur')
            phone = st.text_input('Téléphone')
            submit = st.form_submit_button('Créer')
            if submit:
                if not name.strip():
                    st.error('Nom obligatoire.')
                else:
                    upsert_client({'name': name.strip(), 'sector': sector, 'phone': phone, 'account_type': 'Prospect', 'client_status': 'Actif', 'potential': 'Moyen'})
                    st.success('Client créé.')


def render_sidebar():
    st.sidebar.title('Navigation')
    page = st.sidebar.radio('Aller vers', ['🏠 Accueil', '📂 Affaires', '🏢 Clients', '➕ Ajouter', '📊 Statistiques', '🚨 Aujourd’hui'])
    st.sidebar.markdown('---')
    st.sidebar.caption('MVP local — données stockées en SQLite')
    return page


def apply_style():
    st.set_page_config(page_title='CRM Louis', page_icon='🚛', layout='wide')
    st.markdown(
        '''<style>
        .stMetric {background:#fafafa; padding:12px; border-radius:14px; border:1px solid #eee;}
        .block-container {padding-top: 1rem; padding-bottom: 1rem;}
        div[data-testid="stMetricValue"] {font-size: 1.6rem;}
        .stButton>button {border-radius: 10px;}
        @media (max-width: 768px) {
            .block-container {padding-top: .6rem; padding-left: .7rem; padding-right: .7rem;}
            h1 {font-size: 2rem !important;}
            h2, h3 {font-size: 1.2rem !important;}
            div[data-testid="stMetricValue"] {font-size: 1.25rem !important;}
            div[data-testid="stMetricLabel"] {font-size: .9rem !important;}
            .stButton>button {width: 100%;}
        }
        </style>''',
        unsafe_allow_html=True,
    )


def main():
    apply_style()
    init_db()
    import_seed_if_needed()
    page = render_sidebar()
    if page == '🏠 Accueil':
        page_dashboard()
    elif page == '📂 Affaires':
        page_affairs()
    elif page == '🏢 Clients':
        page_clients()
    elif page == '➕ Ajouter':
        page_quick_add()
    elif page == '📊 Statistiques':
        page_stats()
    else:
        page_today()


if __name__ == '__main__':
    main()
